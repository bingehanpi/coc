import os
import json
from openpyxl import Workbook

def find_latest_cwl_json(prefix="cwl_raw_", suffix=".json"):
    files = [
        f for f in os.listdir(".")
        if f.startswith(prefix) and f.endswith(suffix)
    ]

    if not files:
        raise FileNotFoundError("当前目录下未找到 CWL 原始 JSON 文件")

    files.sort()
    return files[-1]

def load_cwl_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def export_cwl_to_excel(cwl_data, output_file="cwl_raw_readable.xlsx"):
    wb = Workbook()

    # ---------- Sheet 1：Wars ----------
    ws_wars = wb.active
    ws_wars.title = "Wars"
    ws_wars.append([
        "warTag", "state", "teamSize",
        "startTime", "endTime",
        "clanTag", "clanName",
        "opponentTag", "opponentName"
    ])

    # ---------- Sheet 2：Members ----------
    ws_members = wb.create_sheet("Members")
    ws_members.append([
        "warTag", "side",
        "playerTag", "playerName",
        "townHall", "mapPosition"
    ])

    # ---------- Sheet 3：Attacks ----------
    ws_attacks = wb.create_sheet("Attacks")
    ws_attacks.append([
        "warTag", "side",
        "attackerTag", "defenderTag",
        "stars", "destruction",
        "order", "duration"
    ])

    for war_tag, war in cwl_data.items():
        # Wars
        ws_wars.append([
            war_tag,
            war["state"],
            war["teamSize"],
            war["startTime"],
            war["endTime"],
            war["clan"]["tag"],
            war["clan"]["name"],
            war["opponent"]["tag"],
            war["opponent"]["name"]
        ])

        # Members + Attacks
        for side in ("clan", "opponent"):
            for m in war[side]["members"]:
                ws_members.append([
                    war_tag,
                    side,
                    m["tag"],
                    m["name"],
                    m["townhallLevel"],
                    m["mapPosition"]
                ])

                for atk in m.get("attacks", []):
                    ws_attacks.append([
                        war_tag,
                        side,
                        atk["attackerTag"],
                        atk["defenderTag"],
                        atk["stars"],
                        atk["destructionPercentage"],
                        atk["order"],
                        atk.get("duration")
                    ])

    wb.save(output_file)
    print(f"Excel 已生成: {output_file}")

if __name__ == "__main__":
    json_file = find_latest_cwl_json()
    print(f"使用数据文件: {json_file}")

    cwl_data = load_cwl_json(json_file)
    export_cwl_to_excel(cwl_data)
