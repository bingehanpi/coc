from openpyxl import Workbook
import json_excel_raw


def extract_our_clan_data(cwl_data, clan_tag):
    members = []
    attacks = []
    defenses = []
    # 创建标签到玩家名称的映射字典
    player_name_map = {}

    for war_tag, war in cwl_data.items():

        # 关键修复：识别我们在哪一侧
        if war["clan"]["tag"] == clan_tag:
            our_side = "clan"
        elif war["opponent"]["tag"] == clan_tag:
            our_side = "opponent"
        else:
            continue

        clan = war[our_side]

        # 记录我们部落的成员和玩家名称
        for m in clan["members"]:
            player_name_map[m["tag"]] = m["name"]  # 标签 -> 玩家名称映射

            members.append({
                "warTag": war_tag,
                "playerTag": m["tag"],
                "playerName": m["name"],
                "townHall": m["townhallLevel"],
                "mapPosition": m["mapPosition"]
            })

            # 进攻数据（我方打别人）
            for atk in m.get("attacks", []):
                attacker_name = player_name_map.get(atk["attackerTag"], "Unknown")
                defender_name = player_name_map.get(atk["defenderTag"], "Unknown")

                attacks.append({
                    "warTag": war_tag,
                    "attackerTag": attacker_name,  # 用玩家名称代替标签
                    "defenderTag": defender_name,
                    "stars": atk["stars"],
                    "destruction": atk["destructionPercentage"],
                    "order": atk["order"]
                })

            # 防守数据（别人打我）
            if "bestOpponentAttack" in m:
                d = m["bestOpponentAttack"]
                defender_name = player_name_map.get(m["tag"], "Unknown")
                attacker_name = player_name_map.get(d["attackerTag"], "Unknown")

                defenses.append({
                    "warTag": war_tag,
                    "defenderTag": defender_name,  # 用玩家名称代替标签
                    "attackerTag": attacker_name,
                    "stars": d["stars"],
                    "destruction": d["destructionPercentage"]
                })

    return members, attacks, defenses


def export_our_clan_excel(members, attacks, defenses,
                          output_file="cwl_our_clan_only.xlsx"):
    wb = Workbook()

    ws_members = wb.active
    ws_members.title = "Members"
    ws_members.append([
        "warTag", "playerTag", "playerName",
        "townHall", "mapPosition"
    ])

    ws_attacks = wb.create_sheet("Attacks")
    ws_attacks.append([
        "warTag", "attackerName", "defenderName",
        "stars", "destruction", "order"
    ])

    ws_defenses = wb.create_sheet("Defenses")
    ws_defenses.append([
        "warTag", "defenderName", "attackerName",
        "stars", "destruction"
    ])

    # 添加 Members 数据
    for m in members:
        ws_members.append(list(m.values()))

    # 添加 Attacks 数据
    for a in attacks:
        ws_attacks.append(list(a.values()))

    # 添加 Defenses 数据
    for d in defenses:
        ws_defenses.append(list(d.values()))

    wb.save(output_file)
    print(f"已生成我方专用数据文件: {output_file}")

if __name__ == "__main__":
    clan_tag = "#2QL2JYJYC"
    json_file = json_excel_raw.find_latest_cwl_json()
    cwl_data = json_excel_raw.load_cwl_json(json_file)

    members, attacks, defenses = extract_our_clan_data(cwl_data, clan_tag)
    export_our_clan_excel(members, attacks, defenses)
